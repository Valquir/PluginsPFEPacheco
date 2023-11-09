#! python3

# IMPORTS --------------------------------------------------------------------------------------------------------
# Regular
import os, sys
import ctypes
# print("User Current Version:-", sys.version)
sys.path.append(r'C:\Users\vpacheco\AppData\Local\Programs\Python\Python38\Lib\site-packages')
# sys.path.remove(r'C:\Users\vpacheco\AppData\Roaming\pyRevit-Master\site-packages')
# sys.path.remove(r'C:\Users\vpacheco\AppData\Roaming\pyRevit-Master\pyrevitlib')
import pandas as pd
import openpyxl
import xlsxwriter

# .Net Imports
import clr
clr.AddReference("System")
clr.AddReference('RevitAPI')
clr.AddReference('RevitAPIUI')
from System.Collections.Generic import List  # List<elementtype>() <- its special type of list, RevitAPI often requires

# pyRevit + AutoDesk
from Autodesk.Revit.DB import *


# VARIABLES ------------------------------------------------------------------------------------------------------
doc = __revit__.ActiveUIDocument.Document
uidoc = __revit__.ActiveUIDocument
app = __revit__.Application
PATH_SCRIPT = os.path.dirname(__file__)

view_collector = FilteredElementCollector(doc).OfClass(View3D).ToElements()
default_3d_view = next((view for view in view_collector if view.IsTemplate == False), None)

# Phase
all_phases = list(doc.Phases)
phase = all_phases[-1]
category = Category.GetCategory(__revit__.ActiveUIDocument.Document, BuiltInCategory.OST_GenericModel)
category_id = category.Id

# Prendre nom et elevations de chaque niveau
collector_lvl = FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_Levels).WhereElementIsNotElementType().ToElements()
allLevels_elevations = []
allleveles_elev_pieds = []
allLevels_names = []
unitees = {'mm': 1000, 'cm': 100, 'm': 1}
indexunitees = ""
elevacao= 0
NomElevation = ''

for i in collector_lvl:
    allLevels_elevations.append(float(i.get_Parameter(BuiltInParameter.LEVEL_ELEV).AsValueString().replace(',', '.')))
    allLevels_names.append(i.Name)
    allleveles_elev_pieds.append(i.get_BoundingBox(doc.ActiveView).Max.Z - 1.8636)

for i in range(len(allLevels_elevations)):
    if i == 0:
        continue
    else:
        if abs(allLevels_elevations[i]-allLevels_elevations[i-1]) > 900:
            indexunitees = "mm"
        elif abs(allLevels_elevations[i] - allLevels_elevations[i-1]) > 90:
            indexunitees = "cm"
        else:
            indexunitees = "m"

level_elevation = "11111111"       # valeur dummy
level_name = "99999999"            # valeur dummy
level_pieds = "55555555"           # valeur dummy

OrdreAlphabetique = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                     'U', 'V', 'W', 'X', 'Y', 'Z']

# Functions ------------------------------------------------------------------------------
def prendre_volume(element):
    """Function to get respective volume of element.
    :param element : Revit element
    :return : Volume in m3"""

    if element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Escalier" or \
            ((element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Modèles génériques" or
              element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Sols" or
              element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Ossature") and
             element.get_Parameter(BuiltInParameter.HOST_VOLUME_COMPUTED).AsValueString() == "0,00 m³"):

        # Obtenir la boite englobante de le escalier
        bounding_box = element.get_BoundingBox(doc.ActiveView)


        # Calculer le volume de l'escalier
        width = bounding_box.Max.X - bounding_box.Min.X
        length = bounding_box.Max.Y - bounding_box.Min.Y
        height = bounding_box.Max.Z - bounding_box.Min.Z

        if element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Escalier":
            volume = (width * length * height) / (35.315 * 10)

        else:
            volume = (width * length * height) / 35.315

        # Afficher le volume de l'escalier
        return round(volume, 4)

    elif element.get_Parameter(BuiltInParameter.HOST_VOLUME_COMPUTED):
        vol01 = element.get_Parameter(BuiltInParameter.HOST_VOLUME_COMPUTED).AsDouble() / 35.315
        return vol01

    elif element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Rampes d'accès":
        width = float(element.get_Parameter(BuiltInParameter.STAIRS_ATTR_TREAD_WIDTH).AsValueString())
        longueur = float(doc.GetElement(element.GetTypeId()).get_Parameter(BuiltInParameter.RAMP_MAX_RUN_LENGTH).AsValueString())
        epaisseur = float(doc.GetElement(element.GetTypeId()).get_Parameter(BuiltInParameter.RAMP_ATTR_THICKNESS).AsValueString())

        vol3 = width * longueur * epaisseur / unitees[indexunitees]**3

        return vol3

    elif element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Groupes de modèles" or \
         element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Garde-corps" or \
         element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Garde-corps: Mains courantes" or \
         element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Garde-corps: Supports" or \
         element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Garde-corps: Traverses hautes" or \
         element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Escalier: Paliers" or \
         element.get_Parameter(BuiltInParameter.ELEM_CATEGORY_PARAM).AsValueString() == "Escalier: Volées":

        return "N/A"

    else:
        return "ERROR"

def prendre_coordonees(element01):
    """Function to get mid-point of bounding box of element in 3d view.
    :param element01 : Revit element
    :return : mid point (x,y,z)"""
    # Prendre la hauteur selon unités de Revit
    bounding_box = element01.get_BoundingBox(default_3d_view)
    TRF_pointR = bounding_box.Max
    LLB_pointR = bounding_box.Min
    mid_pointR = XYZ((TRF_pointR.X + LLB_pointR.X) / 2,
                     (TRF_pointR.Y + LLB_pointR.Y) / 2,
                     (TRF_pointR.Z + LLB_pointR.Z) / 2)

    MDstring = "("+str(mid_pointR.X)+","+str(mid_pointR.Y)+","+str(mid_pointR.Z)+")"

    return MDstring

def ecrire_excel(filepath, nomFeuille, Colonnes,df):
    """Function write table in Excel."""
    writer = pd.ExcelWriter(filepath, engine='xlsxwriter')
    df.to_excel(writer, sheet_name=nomFeuille)
    workbook = writer.book
    worksheet = writer.sheets[nomFeuille]

    # create a table
    (max_row, max_col) = df.shape
    ZoneTableau = "A1:"+OrdreAlphabetique[max_col]+str(max_row+1)
    worksheet.add_table(ZoneTableau, {"columns": [
                {"header": Colonnes[0]},
                {"header": Colonnes[1]},
                {"header": Colonnes[2]},
                {"header": Colonnes[3]},
                {"header": Colonnes[4]},
                {"header": Colonnes[5]}], })

    # set column width
    worksheet.set_column(0, 0, 9)
    worksheet.set_column(1, 1, 20)
    worksheet.set_column(2, 2, 31)
    worksheet.set_column(4, 4, 20)
    worksheet.set_column(5, 5, 20)

    # save workbook
    writer.close()

def lire_feuille_excel(wb, SheetName):
    dictElements = {}
    sheet_obj = wb[SheetName]
    max_col = sheet_obj.max_column
    max_r = sheet_obj.max_row

    # Will print a particular row value

    for y in range(1, max_r):
        ValeursElements = []
        ElementId = sheet_obj.cell(row=y + 1, column=1).value

        for i in range(2, max_col + 1):
            cell_obj = sheet_obj.cell(row=y + 1, column=i)
            ValeursElements.append(cell_obj.value)

        dictElements[ElementId] = ValeursElements

    return dictElements

